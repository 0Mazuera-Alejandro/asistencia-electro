from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.graphics import Color, Rectangle, RoundedRectangle, Line
from kivy.core.window import Window
from kivy.utils import get_color_from_hex
from kivy.clock import Clock
from kivy.animation import Animation
from datetime import datetime
from database import DatabaseManager
from datetime import datetime

Window.size = (1024, 768)

# ── PALETA DE COLORES ──
C = {
    "fondo":       "#2c3e50",
    "fondo_reg":   "#2d3436",
    "azul":        "#3498db",
    "azul_dark":   "#2980b9",
    "verde":       "#27ae60",
    "verde_dark":  "#1e8449",
    "morado":      "#9b59b6",
    "morado_dark": "#7d3c98",
    "rojo":        "#e74c3c",
    "rojo_dark":   "#c0392b",
    "gris":        "#95a5a6",
    "gris_dark":   "#7f8c8d",
    "blanco":      "#ecf0f1",
    "texto_suave": "#b2bec3",
    "firma":       "#7f8c8d",
}


def hex2kivy(h):
    return get_color_from_hex(h)


def get_jornada():
    hora = datetime.now().hour
    if 6 <= hora < 12:
        return hex2kivy("#F39C12"), "🌅  JORNADA MAÑANA"
    elif 12 <= hora < 18:
        return hex2kivy("#E67E22"), "☀️  JORNADA TARDE"
    elif 18 <= hora <= 21:
        return hex2kivy("#8E44AD"), "🌙  JORNADA NOCHE"
    else:
        return hex2kivy("#E74C3C"), "🔒  LABORATORIO CERRADO"


# ══════════════════════════════════════════
# WIDGETS REUTILIZABLES
# ══════════════════════════════════════════

class FondoColor(BoxLayout):
    """BoxLayout con fondo de color sólido"""
    def __init__(self, color="#2c3e50", radio=0, **kwargs):
        super().__init__(**kwargs)
        self._color = hex2kivy(color)
        self._radio = radio
        with self.canvas.before:
            self._c = Color(*self._color)
            if radio:
                self._rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[radio])
            else:
                self._rect = Rectangle(pos=self.pos, size=self.size)
        self.bind(pos=self._upd, size=self._upd)

    def _upd(self, *a):
        self._rect.pos  = self.pos
        self._rect.size = self.size


class BarraJornada(FondoColor):
    """Barra superior con color e info de jornada"""
    def __init__(self, **kwargs):
        color, texto = get_jornada()
        super().__init__(size_hint_y=None, height=52, **kwargs)
        self._c.rgba = color
        self._color  = color
        self.add_widget(Label(
            text=texto, font_size=17, bold=True,
            color=(1, 1, 1, 1)))


class BtnEstilo(Button):
    """Botón redondeado con colores personalizados y efecto press"""
    def __init__(self, color_normal, color_press, **kwargs):
        super().__init__(**kwargs)
        self.background_color = (0, 0, 0, 0)
        self.color = (1, 1, 1, 1)
        self._cn = hex2kivy(color_normal)
        self._cp = hex2kivy(color_press)
        with self.canvas.before:
            self._bc = Color(*self._cn)
            self._br = RoundedRectangle(pos=self.pos, size=self.size, radius=[14])
        self.bind(pos=self._upd, size=self._upd)

    def _upd(self, *a):
        self._br.pos  = self.pos
        self._br.size = self.size

    def on_press(self):
        self._bc.rgba = self._cp
        anim = Animation(size=(self.width * 0.97, self.height * 0.97),
                         duration=0.08)
        anim.start(self)

    def on_release(self):
        self._bc.rgba = self._cn
        anim = Animation(size=(self.width / 0.97, self.height / 0.97),
                         duration=0.08)
        anim.start(self)


class TarjetaResultado(BoxLayout):
    """Tarjeta con esquinas redondeadas para mostrar resultados"""
    def __init__(self, color, **kwargs):
        super().__init__(orientation="vertical",
                         padding=[25, 18], spacing=10, **kwargs)
        self._color = hex2kivy(color)
        with self.canvas.before:
            self._c = Color(*self._color)
            self._rect = RoundedRectangle(
                pos=self.pos, size=self.size, radius=[16])
        self.bind(pos=self._upd, size=self._upd)

    def _upd(self, *a):
        self._rect.pos  = self.pos
        self._rect.size = self.size


# ══════════════════════════════════════════
# PANTALLA 1: MENÚ PRINCIPAL
# ══════════════════════════════════════════

class MenuPrincipal(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        raiz = FondoColor(color=C["fondo"], orientation="vertical", spacing=0)

        # Barra jornada
        raiz.add_widget(BarraJornada())

        # ── Encabezado ──
        header = FondoColor(color="#243342",
                            orientation="vertical",
                            size_hint_y=None, height=90,
                            padding=[0, 10])
        header.add_widget(Label(
            text="SISTEMA DE ASISTENCIA",
            font_size=30, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=45))
        header.add_widget(Label(
            text="LABORATORIO ELECTRONICA - Universidad del Valle",
            font_size=14, color=hex2kivy(C["texto_suave"]),
            size_hint_y=None, height=25))

        color_jornada, _ = get_jornada()
        linea = FondoColor(size_hint_y=None, height=4)
        linea._c.rgba = color_jornada
        header.add_widget(linea)
        raiz.add_widget(header)

        # ── Logo Univalle centrado ──
        frame_logo = BoxLayout(
            size_hint_y=None, height=160,
            padding=[0, -190])

        try:
            import os
            from kivy.uix.image import Image as KivyImage

            # Ruta absoluta al archivo
            ruta_logo = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "logo_univalle.png")

            if os.path.exists(ruta_logo):
                logo = KivyImage(
                    source=ruta_logo,
                    fit_mode="contain",
                    size_hint=(None, None),
                    size=(280, 350))
                box_centro = BoxLayout()
                box_centro.add_widget(Label())
                box_centro.add_widget(logo)
                box_centro.add_widget(Label())
                frame_logo.add_widget(box_centro)
            else:
                raise FileNotFoundError(f"No existe: {ruta_logo}")

        except Exception as e:
            print(f"Error cargando logo: {e}")
            frame_logo.add_widget(Label(
                text="UNIVERSIDAD DEL VALLE",
                font_size=14, bold=True,
                color=hex2kivy("#e53935")))

        raiz.add_widget(frame_logo)
        # ── Botones ──
        frame = BoxLayout(orientation="vertical",
                          spacing=18, padding=[100, 5, 100, 10])

        botones = [
            ("REGISTRAR ASISTENCIA",        C["azul"],   C["azul_dark"],   "registro_asistencia", 95),
            ("REGISTRAR NUEVO ESTUDIANTE",   C["verde"],  C["verde_dark"],  "registro_estudiante", 95),
            ("REGISTRAR VISITANTE",          C["morado"], C["morado_dark"], "registro_visitante",  95),
            ("PANEL ADMINISTRATIVO",         C["rojo"],   C["rojo_dark"],   "login_admin",         75),
        ]

        for texto, cn, cp, pantalla, alto in botones:
            sombra = FondoColor(color="#000000",
                                size_hint_y=None, height=alto + 4)
            btn = BtnEstilo(
                color_normal=cn, color_press=cp,
                text=texto,
                font_size=20 if alto == 95 else 17,
                bold=True,
                size_hint=(1, 1))
            btn.bind(on_release=lambda x, p=pantalla: self.ir_a(p))
            sombra.add_widget(btn)
            frame.add_widget(sombra)

        raiz.add_widget(frame)

        # Firma
        raiz.add_widget(Label(
            text="V2.2 | Monitor Alejandro Mazuera | Universidad del Valle",
            font_size=10, color=hex2kivy(C["firma"]),
            size_hint_y=None, height=28))

        self.add_widget(raiz)

    def ir_a(self, pantalla):
        self.manager.transition = SlideTransition(direction="left")
        self.manager.current = pantalla

# ══════════════════════════════════════════
# PANTALLA 2: REGISTRO DE ASISTENCIA
# ══════════════════════════════════════════
class RegistroAsistencia(Screen):
    def __init__(self, db, **kwargs):
        super().__init__(**kwargs)
        self.db = db
        self.color_jornada, _ = get_jornada()

        raiz = FondoColor(color=C["fondo_reg"],
                          orientation="vertical", spacing=0)

        # Barra jornada
        raiz.add_widget(BarraJornada())

        # ── Nav ──
        nav = BoxLayout(size_hint_y=None, height=55,
                        padding=[15, 8], spacing=10)
        with nav.canvas.before:
            Color(*hex2kivy("#1e272e"))
            self._nav_rect = Rectangle(pos=nav.pos, size=nav.size)
        nav.bind(pos=lambda i, v: setattr(self._nav_rect, 'pos', v),
                 size=lambda i, v: setattr(self._nav_rect, 'size', v))

        btn_volver = BtnEstilo(
            color_normal=C["gris_dark"], color_press=C["gris"],
            text="←  Volver", font_size=15,
            size_hint=(None, 1), width=130)
        btn_volver.bind(on_release=lambda x: self.volver())
        nav.add_widget(btn_volver)
        nav.add_widget(Label())  # Espaciador
        raiz.add_widget(nav)

        # ── Título ──
        raiz.add_widget(Label(
            text="REGISTRO DE ASISTENCIA",
            font_size=28, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=55))

        # Línea decorativa
        linea = FondoColor(size_hint_y=None, height=4,
                           size_hint_x=None, width=320)
        linea._c.rgba = self.color_jornada
        raiz.add_widget(linea)

        raiz.add_widget(Label(
            text="Escanea o ingresa el código del estudiante",
            font_size=15, color=hex2kivy(C["texto_suave"]),
            size_hint_y=None, height=35))

        # ── Campo entrada con efecto 3D ──
        contenedor_entrada = BoxLayout(
            size_hint_y=None, height=100,
            padding=[180, 8])

        # Marco exterior negro (sombra)
        marco_sombra = FondoColor(color="#000000", radio=0)

        # Marco color jornada
        marco_jornada = BoxLayout(padding=[4, 0, 0, 4])
        with marco_jornada.canvas.before:
            self._mj_color = Color(*self.color_jornada)
            self._mj_rect = Rectangle(pos=marco_jornada.pos,
                                      size=marco_jornada.size)
        marco_jornada.bind(
            pos=lambda i, v: setattr(self._mj_rect, 'pos', v),
            size=lambda i, v: setattr(self._mj_rect, 'size', v))

        # Marco bisel oscuro
        marco_bisel = FondoColor(color="#1a1a1a")

        self.entry_codigo = TextInput(
            hint_text="  Código",
            font_size=26, multiline=False,
            halign="center",
            input_filter='int',          # ← solo números
            background_color=(1, 1, 1, 1),
            foreground_color=hex2kivy("#2d3436"),
            cursor_color=self.color_jornada,
            padding=[10, 15])
        self.entry_codigo.bind(on_text_validate=self.procesar_asistencia)

        marco_bisel.add_widget(self.entry_codigo)
        marco_jornada.add_widget(marco_bisel)
        marco_sombra.add_widget(marco_jornada)
        contenedor_entrada.add_widget(marco_sombra)
        raiz.add_widget(contenedor_entrada)

        raiz.add_widget(Label(
            text="↵  Presiona Enter para confirmar",
            font_size=11, color=hex2kivy("#636e72"),
            size_hint_y=None, height=22))

        # ── Área resultado ──
        self.frame_resultado = BoxLayout(
            orientation="vertical",
            padding=[60, 12], spacing=8)
        raiz.add_widget(self.frame_resultado)

        self.add_widget(raiz)

    def on_enter(self):
        self.entry_codigo.focus = True

    def volver(self):
        self.manager.transition = SlideTransition(direction="right")
        self.manager.current = "menu"

    def procesar_asistencia(self, instance):
        codigo = self.entry_codigo.text.strip()
        if not codigo:
            return

        if codigo.isdigit() and len(codigo) == 7:
            codigo = "20" + codigo

        estudiante = self.db.buscar_estudiante(codigo)

        if estudiante:
            exito, resultado = self.db.registrar_asistencia(codigo)
            if exito:
                self.mostrar_exito(estudiante, resultado)
            else:
                self.mostrar_aviso(resultado)
        else:
            self.mostrar_no_encontrado(codigo)

        self.entry_codigo.text = ""

    def mostrar_exito(self, estudiante, fecha_hora):
        import os
        from kivy.uix.image import Image as KivyImage
        App.get_running_app().reproducir('exito') 
        self.frame_resultado.clear_widgets()
        hora = fecha_hora.split()[1] if " " in fecha_hora else fecha_hora

        # Contenedor principal verde
        contenedor = FondoColor(color="#1e8449", radio=12,
                                orientation="horizontal",
                                padding=[15, 10], spacing=15,
                                size_hint_y=None, height=210)

        # ── IZQUIERDA: datos ──
        col_datos = BoxLayout(orientation="vertical", spacing=6)

        col_datos.add_widget(Label(
            text="ASISTENCIA REGISTRADA",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=32,
            halign="left",
            text_size=(400, None)))

        datos = [
            ("Apellidos:", estudiante['nombre_completo']),
            ("Codigo:",    estudiante['codigo']),
            ("Programa:",  estudiante['nombre_programa'][:30]),
            ("Email:",     estudiante['email'][:35]),
            ("Hora:",      hora),
        ]

        for etiqueta, valor in datos:
            fila = BoxLayout(size_hint_y=None, height=28, spacing=8)
            fila.add_widget(Label(
                text=etiqueta, font_size=13, bold=True,
                color=(1, 1, 1, 1),
                size_hint_x=None, width=90,
                halign="right",
                text_size=(90, None)))
            fila.add_widget(Label(
                text=str(valor), font_size=13,
                color=(1, 1, 1, 1),
                halign="left",
                text_size=(300, None)))
            col_datos.add_widget(fila)

        contenedor.add_widget(col_datos)

        # ── DERECHA: imagen ──
        try:
            ruta_img = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "asistencia_electro.png")

            if os.path.exists(ruta_img):
                img = KivyImage(
                    source=ruta_img,
                    fit_mode="contain",
                    size_hint_x=None,
                    width=180)
                contenedor.add_widget(img)
        except Exception as e:
            print(f"Error imagen: {e}")

        self.frame_resultado.add_widget(contenedor)

        # Animación fade-in
        contenedor.opacity = 0
        Animation(opacity=1, duration=0.5).start(contenedor)

        Clock.schedule_once(
            lambda dt: self.frame_resultado.clear_widgets(), 6)
    def mostrar_aviso(self, mensaje):
        self.frame_resultado.clear_widgets()
        App.get_running_app().reproducir('adver') 
        tarjeta = TarjetaResultado(color="#d35400")
        tarjeta.add_widget(Label(
            text=f"⚠   {mensaje}",
            font_size=17, bold=True, color=(1, 1, 1, 1)))
        self.frame_resultado.add_widget(tarjeta)
        Clock.schedule_once(
            lambda dt: self.frame_resultado.clear_widgets(), 8)

    def mostrar_no_encontrado(self, codigo):
        self.frame_resultado.clear_widgets()
        App.get_running_app().reproducir('error')
        tarjeta = TarjetaResultado(color="#c0392b")

        tarjeta.add_widget(Label(
            text=f"Codigo '{codigo}' no encontrado",
            font_size=18, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=38))

        tarjeta.add_widget(Label(
            text="Por favor registrese primero en el sistema",
            font_size=14, color=(1, 1, 1, 1),
            size_hint_y=None, height=28))

        # Botón registrarse — sin argumento color
        btn_registrar = BtnEstilo(
            color_normal="#ffffff", color_press="#ecf0f1",
            text="Registrarse Ahora",
            font_size=15, bold=True,
            size_hint_y=None, height=50)

        # Color del texto por separado
        btn_registrar.color = hex2kivy("#c0392b")

        btn_registrar.bind(
            on_release=lambda x: self.ir_a_registro(codigo))
        tarjeta.add_widget(btn_registrar)

        self.frame_resultado.add_widget(tarjeta)
        tarjeta.opacity = 0
        Animation(opacity=1, duration=0.4).start(tarjeta)

        Clock.schedule_once(
            lambda dt: self.frame_resultado.clear_widgets(), 10)

    def ir_a_registro(self, codigo):
        self.frame_resultado.clear_widgets()
        self.manager.get_screen(
            "registro_estudiante").codigo_prefill = codigo
        self.manager.transition = SlideTransition(direction="left")
        self.manager.current = "registro_estudiante"


# ══════════════════════════════════════════
# APP PRINCIPAL
# ══════════════════════════════════════════
# ══════════════════════════════════════════
# PANTALLA 3: REGISTRO DE ESTUDIANTE
# ══════════════════════════════════════════
class RegistroEstudiante(Screen):
    def __init__(self, db, **kwargs):
        super().__init__(**kwargs)
        self.db = db
        self.codigo_prefill = ""

        raiz = FondoColor(color=C["fondo"], orientation="vertical", spacing=0)

        # Barra jornada
        raiz.add_widget(BarraJornada())

        # ── Nav ──
        nav = BoxLayout(size_hint_y=None, height=55,
                        padding=[15, 8], spacing=10)
        with nav.canvas.before:
            Color(*hex2kivy("#1e272e"))
            self._nav_rect = Rectangle(pos=nav.pos, size=nav.size)
        nav.bind(pos=lambda i, v: setattr(self._nav_rect, 'pos', v),
                 size=lambda i, v: setattr(self._nav_rect, 'size', v))

        btn_volver = BtnEstilo(
            color_normal=C["gris_dark"], color_press=C["gris"],
            text="←  Volver", font_size=15,
            size_hint=(None, 1), width=130)
        btn_volver.bind(on_release=lambda x: self.volver())
        nav.add_widget(btn_volver)
        nav.add_widget(Label())
        raiz.add_widget(nav)

        # Título
        raiz.add_widget(Label(
            text="REGISTRAR NUEVO ESTUDIANTE",
            font_size=24, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=50))

        # ── Formulario con scroll ──
        scroll = ScrollView(size_hint=(1, 1))
        form = FondoColor(color="#34495e", radio=12,
                          orientation="vertical",
                          spacing=15, padding=[60, 20, 60, 20],
                          size_hint_y=None)
        form.bind(minimum_height=form.setter("height"))

        # Programas disponibles
        self.programas_codigos = {
            "INGENIERÍA ELECTRÓNICA":                 "3744",
            "INGENIERÍA INDUSTRIAL":                  "3751",
            "TECNOLOGÍA EN ELECTRÓNICA INDUSTRIAL":   "2725",
            "TECNOLOGÍA EN ALIMENTOS":                "2712",
            "TECNOLOGÍA DE PROCESAMIENTO DE ALIMENTOS": "2726",
        }

        # ── Campo: Código ──
        form.add_widget(Label(
            text="Código del Estudiante:",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            halign="left", size_hint_y=None, height=30,
            text_size=(Window.width - 120, None)))

        self.entry_codigo = TextInput(
            hint_text="Ej: 202369617",
            font_size=18, multiline=False,
            input_filter='int',          # ← solo números
            size_hint_y=None, height=48,
            background_color=(1, 1, 1, 1),
            foreground_color=hex2kivy("#2c3e50"),
            padding=[12, 10])
        form.add_widget(self.entry_codigo)

        # ── Campo: Programa (botones de selección) ──
        form.add_widget(Label(
            text="Programa Académico:",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            halign="left", size_hint_y=None, height=30,
            text_size=(Window.width - 120, None)))

        self.lbl_programa_seleccionado = Label(
            text="-- Seleccione un programa --",
            font_size=15, color=hex2kivy("#f39c12"),
            size_hint_y=None, height=30,
            halign="left",
            text_size=(Window.width - 120, None))
        form.add_widget(self.lbl_programa_seleccionado)

        self.programa_seleccionado = ""
        self.numero_programa = ""

        grid_programas = GridLayout(cols=1, spacing=8,
                                    size_hint_y=None)
        grid_programas.bind(minimum_height=grid_programas.setter("height"))

        for prog in self.programas_codigos:
            btn = BtnEstilo(
                color_normal=C["azul_dark"], color_press=C["azul"],
                text=prog, font_size=14,
                size_hint_y=None, height=48,
                halign="left", text_size=(Window.width - 150, None))
            btn.bind(on_release=lambda x, p=prog: self.seleccionar_programa(p))
            grid_programas.add_widget(btn)

        form.add_widget(grid_programas)

        # ── Campo: Apellidos ──
        form.add_widget(Label(
            text="Apellidos:",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            halign="left", size_hint_y=None, height=30,
            text_size=(Window.width - 120, None)))

        self.entry_nombre = TextInput(
            hint_text="Ej: LONDOÑO CASTILLO",
            font_size=18, multiline=False,
            size_hint_y=None, height=48,
            background_color=(1, 1, 1, 1),
            foreground_color=hex2kivy("#2c3e50"),
            padding=[12, 10])
        form.add_widget(self.entry_nombre)

        # ── Campo: Email ──
        form.add_widget(Label(
            text="Usuario de Email (sin @):",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            halign="left", size_hint_y=None, height=30,
            text_size=(Window.width - 120, None)))

        email_box = BoxLayout(size_hint_y=None, height=48, spacing=5)
        self.entry_email = TextInput(
            hint_text="usuario",
            font_size=18, multiline=False,
            background_color=(1, 1, 1, 1),
            foreground_color=hex2kivy("#2c3e50"),
            padding=[12, 10])
        email_box.add_widget(self.entry_email)
        email_box.add_widget(Label(
            text="@correounivalle.edu.co",
            font_size=14, color=hex2kivy(C["texto_suave"]),
            size_hint_x=None, width=220))
        form.add_widget(email_box)

        # ── Botones guardar/cancelar ──
        form.add_widget(Label(size_hint_y=None, height=10))

        botones_form = BoxLayout(size_hint_y=None, height=60, spacing=20)

        btn_guardar = BtnEstilo(
            color_normal=C["verde"], color_press=C["verde_dark"],
            text="💾  Guardar", font_size=18, bold=True)
        btn_guardar.bind(on_release=lambda x: self.guardar())
        botones_form.add_widget(btn_guardar)

        btn_cancelar = BtnEstilo(
            color_normal=C["gris"], color_press=C["gris_dark"],
            text="Cancelar", font_size=16)
        btn_cancelar.bind(on_release=lambda x: self.volver())
        botones_form.add_widget(btn_cancelar)

        form.add_widget(botones_form)

        # Mensaje de feedback
        self.lbl_mensaje = Label(
            text="", font_size=15,
            color=hex2kivy("#2ecc71"),
            size_hint_y=None, height=35)
        form.add_widget(self.lbl_mensaje)

        # Frame de éxito — fuera del scroll
        self.frame_exito = BoxLayout(orientation="vertical",
                                      size_hint_y=None, height=0,
                                      opacity=0)
        raiz.add_widget(self.frame_exito)
        scroll.add_widget(form)
        raiz.add_widget(scroll)
        self.add_widget(raiz)

    def on_enter(self):
        # Pre-llenar código si viene de "no encontrado"
        if self.codigo_prefill:
            self.entry_codigo.text = self.codigo_prefill
            self.codigo_prefill = ""
        self.entry_codigo.focus = True

    def seleccionar_programa(self, programa):
        self.programa_seleccionado = programa
        self.numero_programa = self.programas_codigos[programa]
        self.lbl_programa_seleccionado.text = f"✔  {programa} ({self.numero_programa})"
        self.lbl_programa_seleccionado.color = hex2kivy("#2ecc71")

    def guardar(self):
        codigo  = self.entry_codigo.text.strip()
        nombre  = self.entry_nombre.text.strip().upper()
        email   = self.entry_email.text.strip()

        if not codigo:
            self.mostrar_mensaje("Ingrese el codigo del estudiante", error=True)
            return
        # ── Normalizar código ── ← AQUÍ
        if codigo.isdigit() and len(codigo) == 7:
            codigo = "20" + codigo
        if not self.programa_seleccionado:
            self.mostrar_mensaje("Seleccione un programa academico", error=True)
            return
        if not nombre:
            self.mostrar_mensaje("Ingrese los apellidos", error=True)
            return

        if email and '@' not in email:
            email = email + "@correounivalle.edu.co"
        if not email:
            email = f"{codigo}@correounivalle.edu.co"

        exito, mensaje = self.db.agregar_estudiante(
            codigo, self.numero_programa,
            self.programa_seleccionado, nombre, email)

        if exito:
            self.mostrar_panel_exito(codigo, nombre,
                                      self.programa_seleccionado,
                                      self.numero_programa, email)
            # Limpiar formulario
            self.entry_codigo.text = ""
            self.entry_nombre.text = ""
            self.entry_email.text  = ""
            self.programa_seleccionado = ""
            self.numero_programa = ""
            self.lbl_programa_seleccionado.text = "-- Seleccione un programa --"
            self.lbl_programa_seleccionado.color = hex2kivy("#f39c12")
            Clock.schedule_once(lambda dt: self.ocultar_panel_exito(), 6)
            Clock.schedule_once(lambda dt: self.ir_a_asistencia(), 7)
        else:
            self.mostrar_mensaje(f"{mensaje}", error=True)

    def mostrar_panel_exito(self, codigo, nombre, programa,
                             numero_prog, email):
        import os
        from kivy.uix.image import Image as KivyImage

        self.frame_exito.clear_widgets()
        self.frame_exito.height  = 220
        self.frame_exito.opacity = 1

        # Contenedor principal
        contenedor = FondoColor(color="#1e8449", radio=12,
                                orientation="horizontal",
                                padding=[15, 10], spacing=15)

        # ── IZQUIERDA: datos del estudiante ──
        col_datos = BoxLayout(orientation="vertical",
                              spacing=6)

        col_datos.add_widget(Label(
            text="ESTUDIANTE REGISTRADO",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=30,
            halign="left",
            text_size=(400, None)))

        datos = [
            ("Codigo:",    codigo),
            ("Apellidos:", nombre),
            ("Programa:",  programa[:30]),
            ("# Prog:",    numero_prog),
            ("Email:",     email[:35]),
        ]

        for etiqueta, valor in datos:
            fila = BoxLayout(size_hint_y=None, height=26, spacing=8)
            fila.add_widget(Label(
                text=etiqueta, font_size=13, bold=True,
                color=(1, 1, 1, 1),
                size_hint_x=None, width=90,
                halign="right",
                text_size=(90, None)))
            fila.add_widget(Label(
                text=valor, font_size=13,
                color=(1, 1, 1, 1),
                halign="left",
                text_size=(300, None)))
            col_datos.add_widget(fila)

        contenedor.add_widget(col_datos)

        # ── DERECHA: imagen ──
        try:
            ruta_img = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "logo_electro_ardi.png")

            if os.path.exists(ruta_img):
                img = KivyImage(
                    source=ruta_img,
                    fit_mode="contain",
                    size_hint_x=None,
                    width=180)
                contenedor.add_widget(img)
        except Exception as e:
            print(f"Error imagen: {e}")

        self.frame_exito.add_widget(contenedor)

        # Animación fade-in
        self.frame_exito.opacity = 0
        Animation(opacity=1, duration=0.5).start(self.frame_exito)

    def ocultar_panel_exito(self):
        anim = Animation(opacity=0, duration=0.4)
        anim.bind(on_complete=lambda a, w: setattr(
            self.frame_exito, 'height', 0))
        anim.start(self.frame_exito)
    def mostrar_mensaje(self, texto, error=False):
        self.lbl_mensaje.text  = texto
        self.lbl_mensaje.color = hex2kivy("#e74c3c") if error else hex2kivy("#2ecc71")
        Clock.schedule_once(lambda dt: setattr(self.lbl_mensaje, 'text', ''), 4)

    def ir_a_asistencia(self):
        self.manager.transition = SlideTransition(direction="left")
        self.manager.current = "registro_asistencia"

    def volver(self):
        self.manager.transition = SlideTransition(direction="right")
        self.manager.current = "menu"


# ══════════════════════════════════════════
# PANTALLA 4: REGISTRO DE VISITANTE
# ══════════════════════════════════════════
class RegistroVisitante(Screen):
    def __init__(self, db, **kwargs):
        super().__init__(**kwargs)
        self.db = db

        raiz = FondoColor(color=C["fondo"], orientation="vertical", spacing=0)

        # Barra jornada
        raiz.add_widget(BarraJornada())

        # Nav
        nav = BoxLayout(size_hint_y=None, height=55,
                        padding=[15, 8], spacing=10)
        with nav.canvas.before:
            Color(*hex2kivy("#1e272e"))
            self._nav_rect = Rectangle(pos=nav.pos, size=nav.size)
        nav.bind(pos=lambda i, v: setattr(self._nav_rect, 'pos', v),
                 size=lambda i, v: setattr(self._nav_rect, 'size', v))

        btn_volver = BtnEstilo(
            color_normal=C["gris_dark"], color_press=C["gris"],
            text="<-  Volver", font_size=15,
            size_hint=(None, 1), width=130)
        btn_volver.bind(on_release=lambda x: self.volver())
        nav.add_widget(btn_volver)
        nav.add_widget(Label())
        raiz.add_widget(nav)

        # Titulo
        raiz.add_widget(Label(
            text="REGISTRAR VISITANTE",
            font_size=24, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=50))

        # Formulario
        scroll = ScrollView(size_hint=(1, 1))
        form = FondoColor(color="#34495e", radio=12,
                          orientation="vertical",
                          spacing=15, padding=[80, 20, 80, 20],
                          size_hint_y=None)
        form.bind(minimum_height=form.setter("height"))

        # Campo: Nombre completo
        form.add_widget(Label(
            text="Nombre Completo:",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            halign="left", size_hint_y=None, height=30,
            text_size=(Window.width - 160, None)))

        self.entry_nombre = TextInput(
            hint_text="Nombre y apellidos del visitante",
            font_size=18, multiline=False,
            size_hint_y=None, height=48,
            background_color=(1, 1, 1, 1),
            foreground_color=hex2kivy("#2c3e50"),
            padding=[12, 10])
        form.add_widget(self.entry_nombre)

        # Campo: Motivo (fijo, solo lectura visual)
        form.add_widget(Label(
            text="Motivo de la Visita:",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            halign="left", size_hint_y=None, height=30,
            text_size=(Window.width - 160, None)))

        motivo_box = FondoColor(color="#1e8449", radio=8,
                                size_hint_y=None, height=48,
                                padding=[12, 10])
        motivo_box.add_widget(Label(
            text="Visita Interinstitucional",
            font_size=17, bold=True, color=(1, 1, 1, 1),
            halign="left",
            text_size=(Window.width - 200, None)))
        form.add_widget(motivo_box)

        # Campo: Institución
        form.add_widget(Label(
            text="Colegio / Institución:",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            halign="left", size_hint_y=None, height=30,
            text_size=(Window.width - 160, None)))

        self.entry_colegio = TextInput(
            hint_text="Nombre del colegio o institución",
            font_size=18, multiline=False,
            size_hint_y=None, height=48,
            background_color=(1, 1, 1, 1),
            foreground_color=hex2kivy("#2c3e50"),
            padding=[12, 10])
        form.add_widget(self.entry_colegio)

        # Separador
        form.add_widget(Label(size_hint_y=None, height=10))

        # Botones
        botones_form = BoxLayout(size_hint_y=None, height=60, spacing=20)

        btn_guardar = BtnEstilo(
            color_normal=C["verde"], color_press=C["verde_dark"],
            text="Registrar Visitante", font_size=18, bold=True)
        btn_guardar.bind(on_release=lambda x: self.guardar())
        botones_form.add_widget(btn_guardar)

        btn_cancelar = BtnEstilo(
            color_normal=C["gris"], color_press=C["gris_dark"],
            text="Cancelar", font_size=16)
        btn_cancelar.bind(on_release=lambda x: self.volver())
        botones_form.add_widget(btn_cancelar)

        form.add_widget(botones_form)

        # Mensaje feedback
        self.lbl_mensaje = Label(
            text="", font_size=15,
            color=hex2kivy("#2ecc71"),
            size_hint_y=None, height=35)
        form.add_widget(self.lbl_mensaje)

        scroll.add_widget(form)
        raiz.add_widget(scroll)
        self.add_widget(raiz)

    def on_enter(self):
        self.entry_nombre.focus = True

    def guardar(self):
        nombre  = self.entry_nombre.text.strip().upper()
        colegio = self.entry_colegio.text.strip().upper() or "N/A"
        motivo  = "Visita Interinstitucional"

        if not nombre:
            self.mostrar_mensaje("Ingrese el nombre del visitante", error=True)
            return

        exito, resultado = self.db.registrar_visitante(nombre, motivo, colegio)

        if exito:
            hora = resultado.split()[1] if " " in resultado else resultado
            self.mostrar_mensaje(
                f"Visitante registrado  |  Hora: {hora}", error=False)
            self.entry_nombre.text  = ""
            self.entry_colegio.text = ""
            Clock.schedule_once(lambda dt: self.volver(), 2)
        else:
            self.mostrar_mensaje(f"Error: {resultado}", error=True)

    def mostrar_mensaje(self, texto, error=False):
        self.lbl_mensaje.text  = texto
        self.lbl_mensaje.color = (
            hex2kivy("#e74c3c") if error else hex2kivy("#2ecc71"))
        Clock.schedule_once(
            lambda dt: setattr(self.lbl_mensaje, 'text', ''), 4)

    def volver(self):
        self.manager.transition = SlideTransition(direction="right")
        self.manager.current = "menu"



# ══════════════════════════════════════════
# PANTALLA: LOGIN ADMIN
# ══════════════════════════════════════════
class LoginAdmin(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.intentos_fallidos = 0
        raiz = FondoColor(color=C["fondo"], orientation="vertical", spacing=0)

        # Barra jornada
        raiz.add_widget(BarraJornada())

        # Nav
        nav = BoxLayout(size_hint_y=None, height=55,
                        padding=[15, 8], spacing=10)
        with nav.canvas.before:
            Color(*hex2kivy("#1e272e"))
            self._nav_rect = Rectangle(pos=nav.pos, size=nav.size)
        nav.bind(pos=lambda i, v: setattr(self._nav_rect, 'pos', v),
                 size=lambda i, v: setattr(self._nav_rect, 'size', v))

        btn_volver = BtnEstilo(
            color_normal=C["gris_dark"], color_press=C["gris"],
            text="<-  Volver", font_size=15,
            size_hint=(None, 1), width=130)
        btn_volver.bind(on_release=lambda x: self.volver())
        nav.add_widget(btn_volver)
        nav.add_widget(Label())
        raiz.add_widget(nav)

        # Contenido centrado
        contenido = BoxLayout(orientation="vertical",
                              padding=[200, 40], spacing=20)

        contenido.add_widget(Label(
            text="ACCESO ADMINISTRATIVO",
            font_size=24, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=50))

        contenido.add_widget(Label(
            text="Ingrese la contrasena:",
            font_size=16, color=hex2kivy(C["texto_suave"]),
            size_hint_y=None, height=35))

        # Campo contrasena
        self.entry_password = TextInput(
            hint_text="Contrasena",
            font_size=20, multiline=False,
            password=True,              # ← oculta el texto
            size_hint_y=None, height=52,
            background_color=(1, 1, 1, 1),
            foreground_color=hex2kivy("#2c3e50"),
            padding=[12, 12])
        self.entry_password.bind(on_text_validate=self.verificar)
        contenido.add_widget(self.entry_password)

        # Mensaje error
        self.lbl_error = Label(
            text="", font_size=14,
            color=hex2kivy("#e74c3c"),
            size_hint_y=None, height=30)
        contenido.add_widget(self.lbl_error)

        # Botón ingresar
        btn_ingresar = BtnEstilo(
            color_normal=C["rojo"], color_press=C["rojo_dark"],
            text="Ingresar", font_size=18, bold=True,
            size_hint_y=None, height=60)
        btn_ingresar.bind(on_release=lambda x: self.verificar())
        contenido.add_widget(btn_ingresar)

        raiz.add_widget(contenido)
        self.add_widget(raiz)

    def on_enter(self):
        self.entry_password.text = ""
        self.lbl_error.text = ""
        self.entry_password.focus = True

    def verificar(self, *args):
        texto = self.entry_password.text.strip()

        # Palabra de seguridad
        if texto == "S@my":
            self.entry_password.text = ""
            self.lbl_error.text = ""
            self.mostrar_cambio_clave()
            return

        # Verificar contraseña normal
        try:
            with open("admin_password.txt", "r", encoding="utf-8") as f:
                password_correcta = f.read().strip()
        except FileNotFoundError:
            password_correcta = "UNI25LAB"

        if texto == password_correcta:
            self.intentos_fallidos = 0
            self.lbl_error.text = ""
            self.manager.transition = SlideTransition(direction="left")
            self.manager.current = "panel_admin"
        else:
            self.intentos_fallidos += 1
            self.entry_password.text = ""
            self.entry_password.focus = True

            if self.intentos_fallidos >= 2:
                self.lbl_error.text = f"Contrasena incorrecta. Intentos: {self.intentos_fallidos}\nAlerta enviada al administrador."
                # Enviar correo en hilo separado
                import threading
                threading.Thread(
                    target=self.enviar_alerta_email,
                    daemon=True).start()
            else:
                self.lbl_error.text = f"Contrasena incorrecta. Intentos: {self.intentos_fallidos}"

    def enviar_alerta_email(self):
        """Envia alerta por correo de intentos fallidos"""
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        EMAIL_REMITENTE = "laboratorio.electronica.zarzal@correounivalle.edu.co"
        EMAIL_PASSWORD  = "czel bwos petc yogs"
        EMAIL_DESTINO   = "mazuera.manuel@correounivalle.edu.co"

        try:
            msg = MIMEMultipart()
            msg['From']    = EMAIL_REMITENTE
            msg['To']      = EMAIL_DESTINO
            msg['Subject'] = "ALERTA: Intentos fallidos - App Android LAB.ELECTRO"

            fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            cuerpo = f"""
    ALERTA DE SEGURIDAD - APP ANDROID
    ===================================

    Se han detectado multiples intentos fallidos de acceso
    al Panel Administrativo desde la aplicacion Android.

    Detalles:
    - Fecha y Hora: {fecha_hora}
    - Intentos fallidos: {self.intentos_fallidos}
    - Sistema: Laboratorio de Electronica (App Android)

    Por favor, verifique el acceso al sistema.

    ---
    Mensaje automatico del Sistema de Asistencia Android.
            """

            msg.attach(MIMEText(cuerpo, 'plain'))

            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(EMAIL_REMITENTE, EMAIL_PASSWORD)
                smtp.send_message(msg)

            print("Alerta enviada correctamente")

        except Exception as e:
            print(f"Error al enviar alerta: {e}")
    
    def mostrar_cambio_clave(self):
        """Popup para cambiar la contraseña"""
        from kivy.uix.popup import Popup

        # Contenido del popup
        contenido = FondoColor(color="#34495e", orientation="vertical",
                               padding=[20, 15], spacing=12)

        contenido.add_widget(Label(
            text="CAMBIAR CONTRASENA",
            font_size=18, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=35))

        contenido.add_widget(Label(
            text="Nueva contrasena:",
            font_size=14, color=(1, 1, 1, 1),
            size_hint_y=None, height=25,
            halign="left",
            text_size=(400, None)))

        entry_nueva = TextInput(
            hint_text="Nueva contrasena",
            font_size=16, multiline=False,
            password=True,
            size_hint_y=None, height=45,
            background_color=(1, 1, 1, 1),
            foreground_color=hex2kivy("#2c3e50"),
            padding=[10, 10])
        contenido.add_widget(entry_nueva)

        contenido.add_widget(Label(
            text="Confirmar contrasena:",
            font_size=14, color=(1, 1, 1, 1),
            size_hint_y=None, height=25,
            halign="left",
            text_size=(400, None)))

        entry_confirmar = TextInput(
            hint_text="Confirmar contrasena",
            font_size=16, multiline=False,
            password=True,
            size_hint_y=None, height=45,
            background_color=(1, 1, 1, 1),
            foreground_color=hex2kivy("#2c3e50"),
            padding=[10, 10])
        contenido.add_widget(entry_confirmar)

        lbl_msg = Label(
            text="", font_size=13,
            color=hex2kivy("#e74c3c"),
            size_hint_y=None, height=25)
        contenido.add_widget(lbl_msg)

        # Botones
        btns = BoxLayout(size_hint_y=None, height=50, spacing=10)

        popup = Popup(title="",
                      content=contenido,
                      size_hint=(None, None),
                      size=(480, 380),
                      background_color=hex2kivy("#2c3e50"),
                      separator_height=0)

        def guardar(*args):
            nueva    = entry_nueva.text.strip()
            confirmar = entry_confirmar.text.strip()

            if not nueva or not confirmar:
                lbl_msg.text = "Complete ambos campos"
                return
            if nueva != confirmar:
                lbl_msg.text = "Las contrasenas no coinciden"
                return
            if len(nueva) < 6:
                lbl_msg.text = "Minimo 6 caracteres"
                return

            try:
                with open("admin_password.txt", "w", encoding="utf-8") as f:
                    f.write(nueva)
                lbl_msg.color = hex2kivy("#2ecc71")
                lbl_msg.text  = "Contrasena cambiada exitosamente"
                Clock.schedule_once(lambda dt: popup.dismiss(), 1.5)
            except Exception as e:
                lbl_msg.text = f"Error: {e}"

        btn_guardar = BtnEstilo(
            color_normal=C["verde"], color_press=C["verde_dark"],
            text="Guardar", font_size=15, bold=True)
        btn_guardar.bind(on_release=guardar)
        btns.add_widget(btn_guardar)

        btn_cancelar = BtnEstilo(
            color_normal=C["gris"], color_press=C["gris_dark"],
            text="Cancelar", font_size=15)
        btn_cancelar.bind(on_release=lambda x: popup.dismiss())
        btns.add_widget(btn_cancelar)

        contenido.add_widget(btns)
        popup.open()

    def volver(self):
        self.manager.transition = SlideTransition(direction="right")
        self.manager.current = "menu"


# ══════════════════════════════════════════
# PANTALLA 5: PANEL ADMINISTRATIVO
# ══════════════════════════════════════════
class PanelAdmin(Screen):
    def __init__(self, db, **kwargs):
        super().__init__(**kwargs)
        self.db = db
        self.password = "UNI25LAB"

        raiz = FondoColor(color=C["fondo"], orientation="vertical", spacing=0)

        # Barra jornada
        raiz.add_widget(BarraJornada())

        # Nav
        nav = BoxLayout(size_hint_y=None, height=55,
                        padding=[15, 8], spacing=10)
        with nav.canvas.before:
            Color(*hex2kivy("#1e272e"))
            self._nav_rect = Rectangle(pos=nav.pos, size=nav.size)
        nav.bind(pos=lambda i, v: setattr(self._nav_rect, 'pos', v),
                 size=lambda i, v: setattr(self._nav_rect, 'size', v))

        btn_volver = BtnEstilo(
            color_normal=C["rojo_dark"], color_press=C["rojo"],
            text="Cerrar Sesion", font_size=15,
            size_hint=(None, 1), width=160)
        btn_volver.bind(on_release=lambda x: self.volver())
        nav.add_widget(btn_volver)
        nav.add_widget(Label(
            text="PANEL ADMINISTRATIVO",
            font_size=20, bold=True, color=(1, 1, 1, 1)))
        raiz.add_widget(nav)

        # Scroll general
        scroll = ScrollView(size_hint=(1, 1))
        contenido = BoxLayout(orientation="vertical",
                              spacing=15, padding=[20, 15],
                              size_hint_y=None)
        contenido.bind(minimum_height=contenido.setter("height"))

        # Cards de estadísticas
        contenido.add_widget(Label(
            text="ESTADISTICAS EN TIEMPO REAL",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=30))

        self.grid_cards = GridLayout(
            cols=2, spacing=12,
            size_hint_y=None, height=220)
        contenido.add_widget(self.grid_cards)

        # Separador
        contenido.add_widget(Label(
            text="ACCIONES",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=30))

        # Botones de acción
        acciones = [
            ("Ver Estudiantes",        C["azul"],   C["azul_dark"],   self.ver_estudiantes),
            ("Asistencias de Hoy",     C["verde"],  C["verde_dark"],  self.ver_asistencias_hoy),
            ("Todas las Asistencias",  C["morado"], C["morado_dark"], self.ver_asistencias),
            ("Ver Visitantes",         "#6c3483",   "#5b2c6f",        self.ver_visitantes),
            ("Exportar Reporte", "#d35400", "#ba4a00", self.exportar_reporte),
        ]

        grid_acc = GridLayout(cols=2, spacing=12,
                              size_hint_y=None, height=160)
        for texto, cn, cp, comando in acciones:
            btn = BtnEstilo(
                color_normal=cn, color_press=cp,
                text=texto, font_size=16, bold=True,
                size_hint_y=None, height=70)
            btn.bind(on_release=lambda x, c=comando: c())
            grid_acc.add_widget(btn)
        contenido.add_widget(grid_acc)

        # Área de tabla
        contenido.add_widget(Label(
            text="", size_hint_y=None, height=10))

        self.frame_tabla = BoxLayout(
            orientation="vertical",
            size_hint_y=None, spacing=5)
        self.frame_tabla.bind(
            minimum_height=self.frame_tabla.setter("height"))
        contenido.add_widget(self.frame_tabla)

        scroll.add_widget(contenido)
        raiz.add_widget(scroll)
        self.add_widget(raiz)

    def on_enter(self):
        self.cargar_estadisticas()
        
    def exportar_reporte(self):
        """Exporta asistencias como Excel y/o PDF al correo"""
        from kivy.uix.popup import Popup
        from kivy.uix.spinner import Spinner

        contenido = FondoColor(color="#34495e", orientation="vertical",
                               padding=[20, 15], spacing=10)

        contenido.add_widget(Label(
            text="EXPORTAR REPORTE",
            font_size=18, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=35))

        # Años automáticos
        anio_actual = datetime.now().year
        años = [str(a) for a in range(2023, anio_actual + 3)]

        meses = ["Enero", "Febrero", "Marzo", "Abril",
                 "Mayo", "Junio", "Julio", "Agosto",
                 "Septiembre", "Octubre", "Noviembre", "Diciembre"]

        # Año
        fila_anio = BoxLayout(size_hint_y=None, height=44, spacing=10)
        fila_anio.add_widget(Label(
            text="Año:", font_size=14, bold=True,
            color=(1, 1, 1, 1), size_hint_x=None, width=80))
        spinner_anio = Spinner(
            text=str(anio_actual), values=años,
            font_size=15, background_color=hex2kivy(C["azul"]),
            color=(1, 1, 1, 1))
        fila_anio.add_widget(spinner_anio)
        contenido.add_widget(fila_anio)

        # Mes inicio
        fila_ini = BoxLayout(size_hint_y=None, height=44, spacing=10)
        fila_ini.add_widget(Label(
            text="Desde:", font_size=14, bold=True,
            color=(1, 1, 1, 1), size_hint_x=None, width=80))
        spinner_mes_ini = Spinner(
            text="Enero", values=meses,
            font_size=15, background_color=hex2kivy(C["verde_dark"]),
            color=(1, 1, 1, 1))
        fila_ini.add_widget(spinner_mes_ini)
        contenido.add_widget(fila_ini)

        # Mes fin
        fila_fin = BoxLayout(size_hint_y=None, height=44, spacing=10)
        fila_fin.add_widget(Label(
            text="Hasta:", font_size=14, bold=True,
            color=(1, 1, 1, 1), size_hint_x=None, width=80))
        spinner_mes_fin = Spinner(
            text="Diciembre", values=meses,
            font_size=15, background_color=hex2kivy(C["verde_dark"]),
            color=(1, 1, 1, 1))
        fila_fin.add_widget(spinner_mes_fin)
        contenido.add_widget(fila_fin)

        # Mensaje feedback
        lbl_msg = Label(
            text="", font_size=13,
            color=hex2kivy("#2ecc71"),
            size_hint_y=None, height=30)
        contenido.add_widget(lbl_msg)

        # Botones
        btns = BoxLayout(size_hint_y=None, height=55, spacing=10)

        popup = Popup(title="",
                      content=contenido,
                      size_hint=(None, None),
                      size=(460, 360),
                      background_color=hex2kivy("#2c3e50"),
                      separator_height=0)

        def obtener_datos():
            anio    = int(spinner_anio.text)
            mes_ini = meses.index(spinner_mes_ini.text) + 1
            mes_fin = meses.index(spinner_mes_fin.text) + 1

            if mes_ini > mes_fin:
                lbl_msg.color = hex2kivy("#e74c3c")
                lbl_msg.text  = "Mes inicio debe ser menor al mes fin"
                return None, None, None, None

            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT a.fecha_hora, e.codigo,
                       e.nombre_completo, e.nombre_programa,
                       e.numero_programa, a.jornada
                FROM asistencias a
                JOIN estudiantes e ON a.codigo_estudiante = e.codigo
                WHERE strftime('%Y', a.fecha_hora) = ?
                AND CAST(strftime('%m', a.fecha_hora) AS INTEGER) >= ?
                AND CAST(strftime('%m', a.fecha_hora) AS INTEGER) <= ?
                ORDER BY a.fecha_hora ASC
            ''', (str(anio), mes_ini, mes_fin))
            datos = cursor.fetchall()
            conn.close()

            if not datos:
                lbl_msg.color = hex2kivy("#f39c12")
                lbl_msg.text  = "No hay asistencias en ese periodo"
                return None, None, None, None

            nombre = (f"Asistencias_{anio}_"
                      f"{spinner_mes_ini.text[:3]}-"
                      f"{spinner_mes_fin.text[:3]}")
            return datos, anio, mes_ini, nombre

        def enviar_excel(*args):
            import threading

            # Leer valores ANTES del hilo
            anio    = int(spinner_anio.text)
            mes_ini = meses.index(spinner_mes_ini.text) + 1
            mes_fin = meses.index(spinner_mes_fin.text) + 1
            nombre  = (f"Asistencias_{anio}_"
                       f"{spinner_mes_ini.text[:3]}-"
                       f"{spinner_mes_fin.text[:3]}")

            if mes_ini > mes_fin:
                lbl_msg.color = hex2kivy("#e74c3c")
                lbl_msg.text  = "Mes inicio debe ser menor al mes fin"
                return

            lbl_msg.color = hex2kivy("#f39c12")
            lbl_msg.text  = "Generando Excel..."

            threading.Thread(
                target=_enviar_excel_thread,
                args=(anio, mes_ini, mes_fin, nombre),
                daemon=True).start()

        def _enviar_excel_thread(anio, mes_ini, mes_fin, nombre):
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                import io

                # Consultar datos directamente
                conn = self.db.get_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT a.fecha_hora, e.codigo,
                           e.nombre_completo, e.nombre_programa,
                           e.numero_programa, a.jornada
                    FROM asistencias a
                    JOIN estudiantes e ON a.codigo_estudiante = e.codigo
                    WHERE strftime('%Y', a.fecha_hora) = ?
                    AND CAST(strftime('%m', a.fecha_hora) AS INTEGER) >= ?
                    AND CAST(strftime('%m', a.fecha_hora) AS INTEGER) <= ?
                    ORDER BY a.fecha_hora ASC
                ''', (str(anio), mes_ini, mes_fin))
                datos = cursor.fetchall()
                conn.close()

                if not datos:
                    Clock.schedule_once(lambda dt: setattr(
                        lbl_msg, 'color', hex2kivy("#f39c12")), 0)
                    Clock.schedule_once(lambda dt: setattr(
                        lbl_msg, 'text', "No hay asistencias en ese periodo"), 0)
                    return

                # Crear Excel en memoria
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Asistencias"

                encabezados = ["Fecha y Hora", "Codigo",
                               "Apellidos", "Programa",
                               "# Programa", "Jornada"]
                ws.append(encabezados)

                header_fill = PatternFill(start_color="2C3E50",
                                          end_color="2C3E50",
                                          fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                fill_par   = PatternFill(start_color="EBF5FB",
                                         end_color="EBF5FB",
                                         fill_type="solid")
                fill_impar = PatternFill(start_color="FFFFFF",
                                         end_color="FFFFFF",
                                         fill_type="solid")

                for i, fila in enumerate(datos):
                    ws.append(list(fila))
                    fill = fill_par if i % 2 == 0 else fill_impar
                    for cell in ws[i + 2]:
                        cell.fill = fill

                anchos = [20, 14, 28, 38, 12, 12]
                for col, ancho in zip("ABCDEF", anchos):
                    ws.column_dimensions[col].width = ancho

                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                _enviar_correo(
                    buffer.read(),
                    f"{nombre}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    f"Reporte Excel - {nombre}",
                    len(datos))

                Clock.schedule_once(lambda dt: setattr(
                    lbl_msg, 'color', hex2kivy("#2ecc71")), 0)
                Clock.schedule_once(lambda dt: setattr(
                    lbl_msg, 'text',
                    f"Excel enviado ({len(datos)} registros)"), 0)
                Clock.schedule_once(lambda dt: popup.dismiss(), 3)

            except Exception as e:
                Clock.schedule_once(lambda dt: setattr(
                    lbl_msg, 'color', hex2kivy("#e74c3c")), 0)
                Clock.schedule_once(lambda dt: setattr(
                    lbl_msg, 'text', f"Error: {str(e)[:40]}"), 0)

        def enviar_pdf(*args):
            import threading

            # Leer valores ANTES del hilo
            anio    = int(spinner_anio.text)
            mes_ini = meses.index(spinner_mes_ini.text) + 1
            mes_fin = meses.index(spinner_mes_fin.text) + 1
            nombre  = (f"Asistencias_{anio}_"
                       f"{spinner_mes_ini.text[:3]}-"
                       f"{spinner_mes_fin.text[:3]}")

            if mes_ini > mes_fin:
                lbl_msg.color = hex2kivy("#e74c3c")
                lbl_msg.text  = "Mes inicio debe ser menor al mes fin"
                return

            lbl_msg.color = hex2kivy("#f39c12")
            lbl_msg.text  = "Generando PDF..."

            threading.Thread(
                target=_guardar_pdf_thread,
                args=(anio, mes_ini, mes_fin, nombre),
                daemon=True).start()

        def _guardar_pdf_thread(anio, mes_ini, mes_fin, nombre):
            try:
                from reportlab.lib.pagesizes import letter, landscape
                from reportlab.lib import colors
                from reportlab.platypus import (SimpleDocTemplate, Table,
                                                TableStyle, Paragraph,
                                                Spacer, Image)
                from reportlab.lib.styles import getSampleStyleSheet
                import matplotlib
                matplotlib.use('Agg')  # Sin interfaz gráfica
                import matplotlib.pyplot as plt
                import io
                from collections import Counter

                datos = _consultar_datos(anio, mes_ini, mes_fin)

                if not datos:
                    Clock.schedule_once(lambda dt: setattr(
                        lbl_msg, 'color', hex2kivy("#f39c12")), 0)
                    Clock.schedule_once(lambda dt: setattr(
                        lbl_msg, 'text', "No hay asistencias en ese periodo"), 0)
                    return

                ruta  = _ruta_guardado(f"{nombre}.pdf")
                doc   = SimpleDocTemplate(ruta, pagesize=landscape(letter),
                                          leftMargin=30, rightMargin=30,
                                          topMargin=30, bottomMargin=30)
                styles    = getSampleStyleSheet()
                elementos = []

                # ── TÍTULO ──
                elementos.append(Paragraph(
                    "<b>REPORTE DE ASISTENCIAS - LAB.ELECTRO</b>",
                    styles['Title']))
                elementos.append(Paragraph(
                    f"Periodo: {nombre.replace('Asistencias_', '')}  |  "
                    f"Total registros: {len(datos)}  |  "
                    f"Universidad del Valle",
                    styles['Normal']))
                elementos.append(Spacer(1, 15))

                # ── PREPARAR DATOS PARA GRÁFICAS ──

                # 1. Programas
                programas = [d[3] for d in datos]
                conteo_prog = Counter(programas)
                labels_prog  = [p[:30] for p in conteo_prog.keys()]
                valores_prog = list(conteo_prog.values())

                # 2. Jornadas
                jornadas = [d[5] for d in datos]
                conteo_jorn = Counter(jornadas)
                labels_jorn  = list(conteo_jorn.keys())
                valores_jorn = list(conteo_jorn.values())

                # 3. Ingresos por mes
                import calendar
                meses_nombres = [calendar.month_abbr[m] for m in range(1, 13)]
                conteo_mes = Counter()
                for d in datos:
                    try:
                        mes_num = int(str(d[0])[5:7])
                        conteo_mes[mes_num] += 1
                    except:
                        pass
                meses_rango  = list(range(mes_ini, mes_fin + 1))
                valores_mes  = [conteo_mes.get(m, 0) for m in meses_rango]
                labels_mes   = [calendar.month_abbr[m] for m in meses_rango]

                # ── COLORES ──
                colores_pie = ['#3498db', '#e74c3c', '#2ecc71', '#f39c12',
                               '#9b59b6', '#1abc9c', '#e67e22', '#34495e']

                def buf_imagen(fig):
                    buf = io.BytesIO()
                    fig.savefig(buf, format='png', dpi=120,
                                bbox_inches='tight',
                                facecolor='#2c3e50')
                    buf.seek(0)
                    plt.close(fig)
                    return buf

                # ── GRÁFICA 1: Pastel programas ──
                fig1, ax1 = plt.subplots(figsize=(6, 4),
                                          facecolor='#2c3e50')
                ax1.set_facecolor('#2c3e50')
                wedges, texts, autotexts = ax1.pie(
                    valores_prog,
                    labels=None,
                    autopct='%1.1f%%',
                    colors=colores_pie[:len(valores_prog)],
                    startangle=140,
                    wedgeprops={'edgecolor': '#2c3e50', 'linewidth': 2})
                for at in autotexts:
                    at.set_color('white')
                    at.set_fontsize(8)
                ax1.legend(wedges, labels_prog,
                           loc="center left",
                           bbox_to_anchor=(1, 0.5),
                           fontsize=7,
                           labelcolor='white',
                           facecolor='#34495e',
                           edgecolor='#2c3e50')
                ax1.set_title("Asistencias por Programa",
                              color='white', fontsize=11, pad=10)
                buf1 = buf_imagen(fig1)

                # ── GRÁFICA 2: Pastel jornadas ──
                fig2, ax2 = plt.subplots(figsize=(4, 4),
                                          facecolor='#2c3e50')
                ax2.set_facecolor('#2c3e50')
                colores_jorn = {
                    'Mañana': '#F39C12',
                    'Tarde':  '#E67E22',
                    'Noche':  '#8E44AD',
                    'Fuera de horario': '#E74C3C'}
                c_jorn = [colores_jorn.get(j, '#95a5a6')
                          for j in labels_jorn]
                wedges2, texts2, autotexts2 = ax2.pie(
                    valores_jorn,
                    labels=labels_jorn,
                    autopct='%1.1f%%',
                    colors=c_jorn,
                    startangle=90,
                    wedgeprops={'edgecolor': '#2c3e50', 'linewidth': 2})
                for t in texts2:
                    t.set_color('white')
                    t.set_fontsize(9)
                for at in autotexts2:
                    at.set_color('white')
                    at.set_fontsize(8)
                ax2.set_title("Asistencias por Jornada",
                              color='white', fontsize=11, pad=10)
                buf2 = buf_imagen(fig2)

                # ── GRÁFICA 3: Barras por mes ──
                fig3, ax3 = plt.subplots(figsize=(8, 4),
                                          facecolor='#2c3e50')
                ax3.set_facecolor('#34495e')
                barras = ax3.bar(labels_mes, valores_mes,
                                 color='#3498db', edgecolor='#2980b9',
                                 linewidth=1.2)
                ax3.set_title("Ingresos por Mes",
                              color='white', fontsize=11, pad=10)
                ax3.set_xlabel("Mes", color='white', fontsize=9)
                ax3.set_ylabel("Cantidad", color='white', fontsize=9)
                ax3.tick_params(colors='white')
                ax3.spines['bottom'].set_color('#7f8c8d')
                ax3.spines['left'].set_color('#7f8c8d')
                ax3.spines['top'].set_visible(False)
                ax3.spines['right'].set_visible(False)
                ax3.yaxis.grid(True, color='#7f8c8d', alpha=0.4)
                # Valores encima de barras
                for bar in barras:
                    h = bar.get_height()
                    if h > 0:
                        ax3.text(bar.get_x() + bar.get_width() / 2,
                                 h + 0.3, str(int(h)),
                                 ha='center', va='bottom',
                                 color='white', fontsize=8)
                buf3 = buf_imagen(fig3)

                # ── INSERTAR GRÁFICAS EN PDF ──
                # Fila con pastel programas + pastel jornadas
                img1 = Image(buf1, width=310, height=200)
                img2 = Image(buf2, width=210, height=200)

                fila_graficas = Table([[img1, img2]],
                                       colWidths=[320, 220])
                fila_graficas.setStyle(TableStyle([
                    ('ALIGN',   (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN',  (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING',  (0, 0), (-1, -1), 5),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ]))
                elementos.append(fila_graficas)
                elementos.append(Spacer(1, 10))

                # Barras por mes — ancho completo
                img3 = Image(buf3, width=680, height=200)
                elementos.append(img3)
                elementos.append(Spacer(1, 15))

                # ── TABLA DE DATOS ──
                elementos.append(Paragraph(
                    "<b>Detalle de Asistencias</b>", styles['Heading2']))
                elementos.append(Spacer(1, 8))

                encabezados_tabla = [["Fecha/Hora", "Codigo",
                                       "Apellidos", "Programa", "Jornada"]]
                filas_tabla = encabezados_tabla + [
                    [str(d[0])[:16], str(d[1]),
                     str(d[2])[:22], str(d[3])[:28], str(d[5])]
                    for d in datos
                ]

                tabla = Table(filas_tabla, repeatRows=1,
                              colWidths=[100, 70, 140, 200, 70])
                tabla.setStyle(TableStyle([
                    ('BACKGROUND',    (0, 0), (-1, 0),
                     colors.HexColor('#2c3e50')),
                    ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
                    ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE',      (0, 0), (-1, 0), 9),
                    ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                    ('FONTSIZE',      (0, 1), (-1, -1), 7),
                    ('ROWBACKGROUNDS',(0, 1), (-1, -1),
                     [colors.HexColor('#EBF5FB'), colors.white]),
                    ('GRID',          (0, 0), (-1, -1),
                     0.4, colors.HexColor('#bdc3c7')),
                ]))
                elementos.append(tabla)

                doc.build(elementos)

                Clock.schedule_once(lambda dt: setattr(
                    lbl_msg, 'color', hex2kivy("#2ecc71")), 0)
                Clock.schedule_once(lambda dt: setattr(
                    lbl_msg, 'text',
                    f"PDF guardado ({len(datos)} registros)"), 0)
                Clock.schedule_once(lambda dt: popup.dismiss(), 4)

            except Exception as e:
                import traceback
                traceback.print_exc()
                Clock.schedule_once(lambda dt: setattr(
                    lbl_msg, 'color', hex2kivy("#e74c3c")), 0)
                Clock.schedule_once(lambda dt: setattr(
                    lbl_msg, 'text', f"Error: {str(e)[:40]}"), 0)

        def _enviar_correo(archivo_bytes, nombre_archivo,
                   mime_type, asunto, total):
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            from email.mime.base import MIMEBase
            from email import encoders

            EMAIL_REMITENTE = "laboratorio.electronica.zarzal@correounivalle.edu.co"
            EMAIL_PASSWORD  = "czel bwos petc yogs"
            EMAIL_DESTINO   = "mazuera.manuel@correounivalle.edu.co"

            try:
                print(f"[CORREO] Intentando enviar {nombre_archivo}...")
                print(f"[CORREO] Tamaño archivo: {len(archivo_bytes)} bytes")

                msg = MIMEMultipart()
                msg['From']    = EMAIL_REMITENTE
                msg['To']      = EMAIL_DESTINO
                msg['Subject'] = f"Reporte {asunto} - LAB.ELECTRO"

                fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cuerpo = f"""
        Reporte generado desde la App Android del Laboratorio de Electronica.

        Fecha de generacion: {fecha_hora}
        Total de registros: {total}
        Archivo adjunto: {nombre_archivo}

        ---
        Sistema de Asistencia - LAB.ELECTRO
        Universidad del Valle
                """
                msg.attach(MIMEText(cuerpo, 'plain'))

                # Adjuntar archivo
                adjunto = MIMEBase('application', 'octet-stream')
                adjunto.set_payload(archivo_bytes)
                encoders.encode_base64(adjunto)
                adjunto.add_header(
                    'Content-Disposition',
                    f'attachment; filename="{nombre_archivo}"')
                msg.attach(adjunto)

                print("[CORREO] Conectando a Gmail...")
                with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                    print("[CORREO] Haciendo login...")
                    smtp.login(EMAIL_REMITENTE, EMAIL_PASSWORD)
                    print("[CORREO] Enviando mensaje...")
                    smtp.send_message(msg)
                    print("[CORREO] Enviado correctamente!")

            except Exception as e:
                print(f"[CORREO ERROR] {type(e).__name__}: {e}")
                import traceback
                traceback.print_exc()
                # Mostrar error en UI
                Clock.schedule_once(lambda dt: setattr(
                    lbl_msg, 'color', hex2kivy("#e74c3c")), 0)
                Clock.schedule_once(lambda dt: setattr(
                    lbl_msg, 'text', f"Error correo: {str(e)[:45]}"), 0)
        
        # Botones Excel y PDF
        btn_excel = BtnEstilo(
            color_normal=C["verde"], color_press=C["verde_dark"],
            text="Enviar Excel", font_size=14, bold=True)
        btn_excel.bind(on_release=enviar_excel)
        btns.add_widget(btn_excel)

        btn_pdf = BtnEstilo(
            color_normal=C["rojo_dark"], color_press=C["rojo"],
            text="Enviar PDF", font_size=14, bold=True)
        btn_pdf.bind(on_release=enviar_pdf)
        btns.add_widget(btn_pdf)

        btn_cancelar = BtnEstilo(
            color_normal=C["gris"], color_press=C["gris_dark"],
            text="Cancelar", font_size=14)
        btn_cancelar.bind(on_release=lambda x: popup.dismiss())
        btns.add_widget(btn_cancelar)

        contenido.add_widget(btns)
        popup.open()

    def cargar_estadisticas(self):
        self.grid_cards.clear_widgets()
        stats = self.db.obtener_estadisticas()

        cards = [
            ("Estudiantes",      str(stats['total_estudiantes']),  "#2980b9", "#3498db"),
            ("Asistencias Total",str(stats['total_asistencias']),  "#7d3c98", "#9b59b6"),
            ("Asistencias Hoy",  str(stats['asistencias_hoy']),    "#1e8449", "#27ae60"),
            ("Programa Top",
             stats['programa_top'][0][:20],                        "#d35400", "#e67e22"),
        ]

        for titulo, valor, c_top, c_bot in cards:
            card = FondoColor(color=c_bot, radio=12,
                              orientation="vertical",
                              padding=[10, 8], spacing=4,
                              size_hint_y=None, height=100)

            # Franja superior
            franja = FondoColor(color=c_top, radio=0,
                                size_hint_y=None, height=6)
            card.add_widget(franja)

            card.add_widget(Label(
                text=titulo, font_size=13, bold=True,
                color=(1, 1, 1, 1), size_hint_y=None, height=22))
            card.add_widget(Label(
                text=valor, font_size=22, bold=True,
                color=(1, 1, 1, 1), size_hint_y=None, height=35))

            self.grid_cards.add_widget(card)

    def construir_tabla(self, titulo, columnas, datos, fn_refrescar=None):
        self.frame_tabla.clear_widgets()

        # Header con título y botón refrescar
        header_box = BoxLayout(size_hint_y=None, height=40, spacing=10)

        header_box.add_widget(Label(
            text=titulo, font_size=16, bold=True,
            color=(1, 1, 1, 1)))

        if fn_refrescar:
            btn_ref = BtnEstilo(
                color_normal=C["azul_dark"], color_press=C["azul"],
                text="Refrescar", font_size=13,
                size_hint=(None, 1), width=120)
            btn_ref.bind(on_release=lambda x: fn_refrescar())
            header_box.add_widget(btn_ref)

        self.frame_tabla.add_widget(header_box)

        if not datos:
            self.frame_tabla.add_widget(Label(
                text="No hay datos disponibles",
                font_size=14, color=hex2kivy(C["texto_suave"]),
                size_hint_y=None, height=40))
            return

        # Encabezados
        header = GridLayout(cols=len(columnas),
                            size_hint_y=None, height=38)
        with header.canvas.before:
            Color(*hex2kivy("#2c3e50"))
            self._h_rect = Rectangle(pos=header.pos, size=header.size)
        header.bind(
            pos=lambda i, v: setattr(self._h_rect, 'pos', v),
            size=lambda i, v: setattr(self._h_rect, 'size', v))

        for col in columnas:
            header.add_widget(Label(
                text=col, font_size=13, bold=True,
                color=(1, 1, 1, 1)))
        self.frame_tabla.add_widget(header)

        # Filas
        for i, fila in enumerate(datos):
            color_fila = "#34495e" if i % 2 == 0 else "#2c3e50"
            row = GridLayout(cols=len(columnas),
                             size_hint_y=None, height=36)
            with row.canvas.before:
                Color(*hex2kivy(color_fila))
                _r = Rectangle(pos=row.pos, size=row.size)
            row.bind(
                pos=lambda i2, v, r=_r: setattr(r, 'pos', v),
                size=lambda i2, v, r=_r: setattr(r, 'size', v))

            for celda in fila:
                row.add_widget(Label(
                    text=str(celda)[:25],
                    font_size=12, color=(1, 1, 1, 1)))
            self.frame_tabla.add_widget(row)
    

    def ver_estudiantes(self):
        self.cargar_estadisticas()
        datos = self.db.obtener_todos_estudiantes()
        datos_fmt = [(str(d[0]), str(d[2])[:20], str(d[3])[:20])
                     for d in datos]
        self.construir_tabla_estudiantes(datos_fmt)

    def construir_tabla_estudiantes(self, datos):
        """Tabla especial de estudiantes con botón eliminar"""
        self.frame_tabla.clear_widgets()

        # Header
        header_box = BoxLayout(size_hint_y=None, height=40, spacing=10)
        header_box.add_widget(Label(
            text=f"ESTUDIANTES ({len(datos)})",
            font_size=16, bold=True, color=(1, 1, 1, 1)))

        btn_ref = BtnEstilo(
            color_normal=C["azul_dark"], color_press=C["azul"],
            text="Refrescar", font_size=13,
            size_hint=(None, 1), width=120)
        btn_ref.bind(on_release=lambda x: self.ver_estudiantes())
        header_box.add_widget(btn_ref)
        self.frame_tabla.add_widget(header_box)

        if not datos:
            self.frame_tabla.add_widget(Label(
                text="No hay estudiantes registrados",
                font_size=14, color=hex2kivy(C["texto_suave"]),
                size_hint_y=None, height=40))
            return

        # Encabezados
        header = GridLayout(cols=4, size_hint_y=None, height=38)
        with header.canvas.before:
            Color(*hex2kivy("#2c3e50"))
            self._h_rect = Rectangle(pos=header.pos, size=header.size)
        header.bind(
            pos=lambda i, v: setattr(self._h_rect, 'pos', v),
            size=lambda i, v: setattr(self._h_rect, 'size', v))

        for col in ["Codigo", "Programa", "Apellidos", "Accion"]:
            header.add_widget(Label(
                text=col, font_size=13, bold=True,
                color=(1, 1, 1, 1)))
        self.frame_tabla.add_widget(header)

        # Filas con botón eliminar
        for i, fila in enumerate(datos):
            color_fila = "#34495e" if i % 2 == 0 else "#2c3e50"
            row = GridLayout(cols=4, size_hint_y=None, height=42)
            with row.canvas.before:
                Color(*hex2kivy(color_fila))
                _r = Rectangle(pos=row.pos, size=row.size)
            row.bind(
                pos=lambda i2, v, r=_r: setattr(r, 'pos', v),
                size=lambda i2, v, r=_r: setattr(r, 'size', v))

            for celda in fila:
                row.add_widget(Label(
                    text=str(celda)[:22],
                    font_size=12, color=(1, 1, 1, 1)))

            # Botón eliminar
            btn_del = BtnEstilo(
                color_normal=C["rojo_dark"], color_press=C["rojo"],
                text="Eliminar", font_size=12,
                size_hint_y=None, height=34)
            btn_del.bind(
                on_release=lambda x, cod=fila[0]: self.confirmar_eliminar(cod))
            row.add_widget(btn_del)

            self.frame_tabla.add_widget(row)

    def confirmar_eliminar(self, codigo):
        """Popup de confirmación antes de eliminar"""
        from kivy.uix.popup import Popup

        contenido = FondoColor(color="#34495e", orientation="vertical",
                               padding=[20, 15], spacing=15)

        contenido.add_widget(Label(
            text=f"Eliminar estudiante\n{codigo}?",
            font_size=17, bold=True, color=(1, 1, 1, 1),
            halign="center", size_hint_y=None, height=60))

        contenido.add_widget(Label(
            text="Esto eliminara tambien\nsus asistencias registradas.",
            font_size=14, color=hex2kivy("#f39c12"),
            halign="center", size_hint_y=None, height=50))

        btns = BoxLayout(size_hint_y=None, height=55, spacing=15)

        popup = Popup(title="",
                      content=contenido,
                      size_hint=(None, None),
                      size=(400, 260),
                      background_color=hex2kivy("#2c3e50"),
                      separator_height=0)

        def eliminar(*args):
            exito, mensaje = self.db.eliminar_estudiante(codigo)
            popup.dismiss()
            if exito:
                self.ver_estudiantes()  # Refrescar tabla
                self.cargar_estadisticas()
            else:
                self.lbl_error_temp(mensaje)

        btn_si = BtnEstilo(
            color_normal=C["rojo"], color_press=C["rojo_dark"],
            text="Si, eliminar", font_size=15, bold=True)
        btn_si.bind(on_release=eliminar)
        btns.add_widget(btn_si)

        btn_no = BtnEstilo(
            color_normal=C["gris"], color_press=C["gris_dark"],
            text="Cancelar", font_size=15)
        btn_no.bind(on_release=lambda x: popup.dismiss())
        btns.add_widget(btn_no)

        contenido.add_widget(btns)
        popup.open()

    def lbl_error_temp(self, mensaje):
        """Muestra mensaje temporal de error"""
        lbl = Label(
            text=mensaje, font_size=14,
            color=hex2kivy("#e74c3c"),
            size_hint_y=None, height=35)
        self.frame_tabla.add_widget(lbl)
        Clock.schedule_once(
            lambda dt: self.frame_tabla.remove_widget(lbl), 3)

    def ver_asistencias_hoy(self):
        self.cargar_estadisticas()
        self.frame_tabla.clear_widgets()

        # Título
        self.frame_tabla.add_widget(Label(
            text="ASISTENCIAS DE HOY",
            font_size=16, bold=True, color=(1, 1, 1, 1),
            size_hint_y=None, height=35))

        # Botones de jornada
        botones_jornada = GridLayout(cols=5, spacing=8,
                                      size_hint_y=None, height=55)

        jornadas = [
            ("Todas",   "#27ae60", "#1e8449", None),
            ("Manana",  "#f39c12", "#d68910", "Mañana"),
            ("Tarde",   "#e67e22", "#ca6f1e", "Tarde"),
            ("Noche",   "#8e44ad", "#7d3c98", "Noche"),
            ("Refrescar", C["azul_dark"], C["azul"], "refresh"),
        ]

        for texto, cn, cp, filtro in jornadas:
            btn = BtnEstilo(
                color_normal=cn, color_press=cp,
                text=texto, font_size=13, bold=True)
            btn.bind(on_release=lambda x, f=filtro: self.cargar_tabla_hoy(f))
            botones_jornada.add_widget(btn)

        self.frame_tabla.add_widget(botones_jornada)

        # Cargar todas por defecto
        self.cargar_tabla_hoy(None)

    def cargar_tabla_hoy(self, jornada):
        """Carga asistencias de hoy filtradas por jornada"""
        if jornada == "refresh":
            jornada = None

        hoy = datetime.now().strftime("%Y-%m-%d")
        conn = self.db.get_connection()
        cursor = conn.cursor()

        if jornada:
            cursor.execute('''
                SELECT time(a.fecha_hora), e.codigo,
                       e.nombre_completo, a.jornada
                FROM asistencias a
                JOIN estudiantes e ON a.codigo_estudiante = e.codigo
                WHERE date(a.fecha_hora) = ?
                AND a.jornada = ?
                ORDER BY a.fecha_hora DESC
            ''', (hoy, jornada))
        else:
            cursor.execute('''
                SELECT time(a.fecha_hora), e.codigo,
                       e.nombre_completo, a.jornada
                FROM asistencias a
                JOIN estudiantes e ON a.codigo_estudiante = e.codigo
                WHERE date(a.fecha_hora) = ?
                ORDER BY a.fecha_hora DESC
            ''', (hoy,))

        datos = cursor.fetchall()
        conn.close()

        # ── Limpiar TODO lo que está debajo de los botones ──
        # Los botones son el índice -1 (último agregado = primero en children)
        # Conservar solo título (index -2) y botones (index -1)
        while len(self.frame_tabla.children) > 2:
            self.frame_tabla.remove_widget(
                self.frame_tabla.children[0])

        colores_jornada = {
            "Mañana": "#7b6000",
            "Tarde":  "#7b3300",
            "Noche":  "#1a237e",
        }

        if not datos:
            self.frame_tabla.add_widget(Label(
                text="No hay asistencias para esta jornada",
                font_size=14, color=hex2kivy(C["texto_suave"]),
                size_hint_y=None, height=40))
            return

        # Encabezados
        header = GridLayout(cols=4, size_hint_y=None, height=38)
        with header.canvas.before:
            Color(*hex2kivy("#2c3e50"))
            self._h_rect2 = Rectangle(pos=header.pos, size=header.size)
        header.bind(
            pos=lambda i, v: setattr(self._h_rect2, 'pos', v),
            size=lambda i, v: setattr(self._h_rect2, 'size', v))

        for col in ["Hora", "Codigo", "Apellidos", "Jornada"]:
            header.add_widget(Label(
                text=col, font_size=13, bold=True,
                color=(1, 1, 1, 1)))
        self.frame_tabla.add_widget(header)

        # Filas
        for i, fila in enumerate(datos):
            jornada_fila = fila[3]
            color_fila = "#34495e" if i % 2 == 0 else "#2c3e50"

            row = GridLayout(cols=4, size_hint_y=None, height=38)
            with row.canvas.before:
                Color(*hex2kivy(color_fila))
                _r = Rectangle(pos=row.pos, size=row.size)
            row.bind(
                pos=lambda i2, v, r=_r: setattr(r, 'pos', v),
                size=lambda i2, v, r=_r: setattr(r, 'size', v))

            for celda in fila[:3]:
                row.add_widget(Label(
                    text=str(celda)[:22],
                    font_size=12, color=(1, 1, 1, 1)))

            color_badge = colores_jornada.get(jornada_fila, "#2c3e50")
            badge = FondoColor(color=color_badge, radio=8, padding=[4, 4])
            badge.add_widget(Label(
                text=jornada_fila, font_size=11,
                bold=True, color=(1, 1, 1, 1)))
            row.add_widget(badge)

            self.frame_tabla.add_widget(row)
    
    def ver_asistencias(self):
        self.cargar_estadisticas()
        datos = self.db.obtener_todas_asistencias()
        datos_fmt = [(str(d[1])[:16], str(d[2]), str(d[3])[:18])
                     for d in datos]
        self.construir_tabla(
            f"TODAS LAS ASISTENCIAS ({len(datos_fmt)})",
            ["Fecha/Hora", "Codigo", "Nombre"],
            datos_fmt,
            self.ver_asistencias)

    def ver_visitantes(self):
        self.cargar_estadisticas()
        datos = self.db.obtener_todos_visitantes()
        datos_fmt = [(str(d[1])[:16], str(d[2])[:20], str(d[4])[:20])
                     for d in datos]
        self.construir_tabla(
            f"VISITANTES ({len(datos_fmt)})",
            ["Fecha/Hora", "Nombre", "Institucion"],
            datos_fmt,
            self.ver_visitantes)

    def volver(self):
        self.manager.transition = SlideTransition(direction="right")
        self.manager.current = "menu"

class AsistenciaApp(App):
    def build(self):
        self.title = "Sistema de Asistencia - LAB.ELECTRO"
        self.db = DatabaseManager()

        # ── Inicializar sonidos ──
        try:
            import pygame
            pygame.mixer.init()
            import os
            base = os.path.dirname(os.path.abspath(__file__))
            self.sonido_exito = pygame.mixer.Sound(os.path.join(base, "exito.mp3"))
            self.sonido_error = pygame.mixer.Sound(os.path.join(base, "error.mp3"))
            self.sonido_adver = pygame.mixer.Sound(os.path.join(base, "adver.mp3"))
            self.sonidos_ok = True
        except Exception as e:
            print(f"Sonidos no disponibles: {e}")
            self.sonidos_ok = False

        sm = ScreenManager()
        sm.add_widget(MenuPrincipal(name="menu"))
        sm.add_widget(RegistroAsistencia(db=self.db, name="registro_asistencia"))
        sm.add_widget(RegistroEstudiante(db=self.db, name="registro_estudiante"))
        sm.add_widget(RegistroVisitante(db=self.db, name="registro_visitante"))
        sm.add_widget(PanelAdmin(db=self.db, name="panel_admin"))
        sm.add_widget(LoginAdmin(name="login_admin"))
        return sm

    def reproducir(self, tipo):
        """tipo: 'exito', 'error', 'adver'"""
        if not self.sonidos_ok:
            return
        try:
            import pygame
            sonidos = {
                'exito': self.sonido_exito,
                'error': self.sonido_error,
                'adver': self.sonido_adver,
            }
            s = sonidos.get(tipo)
            if s:
                s.play()
        except Exception as e:
            print(f"Error reproduciendo sonido: {e}")


if __name__ == "__main__":
    AsistenciaApp().run()